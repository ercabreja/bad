from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import select
import openpyxl

# abrir el archivo de Excel
wb = openpyxl.load_workbook("C:\Python-Selenium\PruebasQAutomation\CargaDatosExcellaGooForm\datos.xlsx")

#obtener la primera hoja
hoja1 = wb.worksheets[0]

#inicilizar el controlador de selenium
driver = webdriver.Firefox(executable_path="C:\Python-Selenium\PruebasQAutomation\firefox_webdriver\geckodriver.exe")


# abrir pagina del formulario y iterar sobre todos los valores en la hoja de calculo

for fila in range(2, 7):
    driver.get("https://forms.gle/b2p7JKXTELwf9BLTA")
    time.sleep(1)
    nombre = hoja1.cell(row=fila, column=1).value
    edad = hoja1.cell(row=fila, column=2).value
    nivel_idiomas = hoja1.cell(row=fila, column=3).value


    #Capturar y llenar los campos del formulario
    driver.find_element (By.XPATH, value="(//input[@type='text'])[1]").send_keys(nombre)
    driver.find_element (By.XPATH, value="(//input[@type='text'])[2]").send_keys(edad)
    
    if (nivel_idiomas) == "Alto":
            driver.find_element (By.XPATH, value="(//span[normalize-space()='Alto'])[1]").click()
    
    elif (nivel_idiomas) == "Medio":
            driver.find_element (By.XPATH, value="(//span[normalize-space()='Medio'])[1]").click()

    elif (nivel_idiomas) == "Bajo":
            driver.find_element (By.XPATH, value="(//span[normalize-space()='Bajo'])[1]").click()
    
    else:
            driver.find_element (By.XPATH, value="//span[@dir='auto'][contains(.,'N/A')]").click()


    driver.find_element (By.XPATH, value="(//span[contains(text(),'Enviar')])[1]").click()
    time.sleep(2)
driver.close()